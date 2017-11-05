#!/usr/bin/env python
import openpyxl

""" Goes through the production plan excel sheet and returns a list of products"""
def load_production_plan():
    more_products = True
    incrementer = 3

    production_list = []

    # Adds all the cells that have a value to the production list
    while more_products:
        if sheet_one.cell(row=incrementer, column=1).value:
            production_list.append(sheet_one.cell(row=incrementer, column=1).value)
        else:
            more_products = False

        incrementer += 1

    return production_list

# Uses the openpyxl library to open the production plan excel file 
workbook = openpyxl.load_workbook('/home/sev-ros/catkin_ws/src/beginner_tutorials/scripts/Daily_Production_Plan.xlsx')
type(workbook)

# Stores the first sheet of the excel file
sheet_one = workbook.get_sheet_by_name('Plan')

# If this is run as main the cells will be printed
if __name__ == "__main__":
    more_to_print = True
    i = 3

    while more_to_print:
        if sheet_one.cell(row=i, column=1).value:
            print(sheet_one.cell(row=i, column=1).value)
        else:
            more_to_print = False

        i += 1
